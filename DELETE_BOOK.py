import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import messagebox

def delete():      #delete function 
    df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #reading from the excel file
    bookid= bookid_entry.get()     #variable for book ID
    Bookname=Bookname_entry.get()   #variable for Book Name
    if (bookid=="" or Bookname==" "):
        messagebox.showwarning("Invalid","All the fields are required")
        return
    else:
        for x in df.index:       
            if df['bookid'][x]==bookid:    #condition checking if the book is present or not
                msg=messagebox.showinfo("BOOK IS PRESENT","CAN BE DELETED")
                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #loading the workbook
                sheet = wb.active     #making the excel file actie to work on
                a=df['serial'][x]         
                b=a+1              #logic for rows so that it gets incremented accordingly
                if(sheet.cell(row=b, column=6).value == bookid):    # checking if the bookid entered is present in bookid column
                    sheet.cell(row=b, column=1).value = "-"        #if does then replace it with "-"
                    sheet.cell(row=b, column=2).value = "-"
                    sheet.cell(row=b, column=3).value = "-"
                    sheet.cell(row=b, column=4).value = "-"
                    sheet.cell(row=b, column=5).value = "-"
                    sheet.cell(row=b, column=6).value = "-"
                    sheet.cell(row=b, column=7).value = "-"
                    sheet.cell(row=b, column=8).value = "-"
                    sheet.cell(row=b, column=9).value = "-"
                    sheet.cell(row=b, column=10).value = "-"
                    sheet.cell(row=b, column=11).value = "-"
                    sheet.cell(row=b, column=12).value = "-"  
                    
                    index_row = []
                    for i in range(1, sheet.max_row):          #for shifting the next row in place of deleted or made "-" row
                        if sheet.cell(i, 1).value == "-":
                            index_row.append(i)
                    for row_del in range(len(index_row)):
                        sheet.delete_rows(idx=index_row[row_del], amount=1)
                        index_row = list(map(lambda k: k - 1, index_row))
                    
                    #wb.save(r"C:\Users\Admin\Downloads\export_issue2.xlsx")   #saving the excel file after execution
                    wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx") 
                break
        if df['bookid'][x]==bookid:
                #msg=messagebox.showinfo("Success","Book have been deleted successfully")
                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #loading the workbook
                sheet = wb.active     
                counter = 0
                for column in range(1,2):
                    column_letter = get_column_letter(column)
                for row in range(2,sheet.max_row+1):
                    counter = counter +1
                    sheet[column_letter + str(row)] = counter
                wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx.xlsx") 
                msg=messagebox.showinfo("Success","Book have been deleted successfully")
                Bookname_entry.delete(0, END)
                bookid_entry.delete(0, END)
        else:
                msg=messagebox.showinfo("Book Not Present","Cannot be deleted") 
def Delete_book():     #the main function which is called upon clicking the delete button in the menu
    root = Tk()
    root.title("Library System")
    root.geometry("800x450")
    frame1 = Frame(root,bg="#fafafa")
    frame1.pack(expand=True,fill=BOTH)
    
    headingFrame12 = Frame(frame1,bg="#d91a40",bd=5) 
    headingFrame12.place(relx=0.22,rely=0.025,relwidth=0.55,relheight=0.10)
    headingLabel2 = Label(headingFrame12, text="Delete a Book", bg="#fc9a44", fg='white', font=('Courier',19,'bold'))
    headingLabel2.place(relx=0,rely=0, relwidth=1, relheight=1)
    bookid= StringVar()
    Bookname=StringVar()
    global bookid_entry,Bookname_entry
    
    frame2 = Frame(frame1 ,bg="#fc9a44") 
    frame2.place(relx=0.05,rely=0.20,relwidth=0.9,relheight=0.6)
    
    lb_Bookname=Label(frame2,text="Enter the book name to be deleted",font=("calibre",13,"normal"),bg="#fc9a44")
    lb_Bookname.place(relx=0.05,rely=0.30, relheight=0.08)
   
    Bookname_entry=Entry(frame2,textvariable=Bookname,font=('calibre',13,'bold'))
    Bookname_entry.place(relx=0.42,rely=0.30, relwidth=0.50, relheight=0.09)
        
    lb_bookid=Label(frame2,text="Enter the Book ID",font=("calibre",13,"normal"),bg="#fc9a44")
    lb_bookid.place(relx=0.05,rely=0.60, relheight=0.08)
    
    bookid_entry=Entry(frame2,textvariable=bookid,font=('calibre',13,'bold'))
    bookid_entry.place(relx=0.42,rely=0.60, relwidth=0.50, relheight=0.09)
    
    sub_button=Button(frame1,text="Delete",command= delete,bg='#d4d4d4', fg='black',font=("calibre",12,'normal'))
    sub_button.place(relx=0.40,rely=0.85, relwidth=0.18,relheight=0.08)


    root.mainloop()
def main():
    Delete_book()
if __name__=='__main__':
    main()