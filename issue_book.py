import openpyxl as op
from openpyxl import Workbook

import datetime as dt
import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import messagebox
root = Tk()
root.geometry("400x400")
frame1 = Frame(root, height=30 , background="bisque")
frame1.pack_propagate(False)
frame1.pack(fill=BOTH,side = TOP)
def onClick1():

    frame2 = Frame(root , width=60, height=200 , background="red")
    frame2.pack_propagate(False)
#the main function for isuing a book 1-it will search bookid,2-it will check availablity,3-it will issue book and open the workbook and save all the details

   # the issue function
    def availissuebook():
            bookid=bookid_entry.get()
            issuername=issuername_entry.get()
            rollnumber=roll_entry.get()
            issuingdate=issuedate_entry.get()
            expiry=expiry_entry.get()

            nospace=0
            df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
            for x in df.index:
                if df['bookid'][x]==bookid:
                        if df['issuername'][x]=="none" or df['issuername'][x]=='NONE':
                            nospace=1
                            break
                        else :
                            continue

            if(nospace==1):
                    msg = messagebox.showinfo("can be issued","available")
                    df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")

                    counter = 0
                    for x in df['bookid']:

                        counter = counter + 1

                        #if df['bookid'][x]==bookid:
                        if x == bookid:
                            #opening the excel sheet using openpyxl
                            print("counter is"+str(counter))
                            wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                            sheet = wb.active
                            i=counter +1
                            sheet.cell(row=i, column=7).value = issuername
                            print(sheet.cell(row=i, column=7).value)


                            sheet.cell(row=i,column=8).value=rollnumber
                            sheet.cell(row=i,column=9).value=issuingdate

                            sheet.cell(row=i,column=10).value=expiry

                            counter=0



                            wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                            msg=messagebox.showinfo("success","book is issued")#/*
                             #this will delete the content given on screen for refilling it

                            bookid_entry.delete(0, END)
                            issuername_entry.delete(0, END)
                            roll_entry.delete(0, END)
                            issuedate_entry.delete(0, END)
                            expiry_entry.delete(0, END)
                            messagebox.showinfo("want to issue books?","Enter the data in the blank space and press 'Submit' or press 'Quit' to exit")

                            break
            else:
                    msg = messagebox.showinfo("cant be issued","no availability")
    #calling search function to search if book id is available,then if book found pass it to the issue book if available function
    #if book is not present pop up will come and book is unavailable to be issued or book not found

    def searching():
            bookid=bookid_entry.get()
            df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
            found=0
            for x in df.index:
                if df['bookid'][x]==str(bookid):
                     found=1
                else:
                    continue
            if (found==1):
                msg = messagebox.showinfo("search complted","book  found")
                s2_button=tk.Button(frame3, text="issuebook if AV",command= availissuebook)
                s2_button.grid(row=0,column=13)
            else:
                msg = messagebox.showinfo("search complted","book not found")

    frame3 = Frame(root , width=400, height=500 , background="pink")
    frame3.pack_propagate(False)
    frame3.pack(side=TOP,fill=BOTH)

    compare=tk.StringVar()
    bookid=tk.StringVar()
    issuername=tk.StringVar()
    rollnumber=tk.StringVar()
    issuingdate=tk.StringVar()
    expiry=tk.StringVar()
    lb_none=tk.Label(frame3,text="no",width=1,font=("calibre",10,"normal"),bg="white")
    lb_none.grid(row=0,column=23)

    lb_bookid=tk.Label(frame3,text="enter the book to be searched",width=70,font=("calibre",13,"normal"),bg="white")
    lb_bookid.grid(row=0,column=0)
    bookid_entry=tk.Entry(frame3,textvariable=bookid,width=20,font=('calibre',13,'bold'))
    bookid_entry.grid(row=0,column=1)

    sub_button=tk.Button(frame3, text="search",command= searching)
    sub_button.grid(row=0,column=14)

    lb_issuername=tk.Label(frame3,text="enter the name of issuer",width=70,font=("calibre",13,"normal"),bg="white")
    lb_issuername.grid(row=1,column=0)
    issuername_entry=tk.Entry(frame3,textvariable=issuername,width=20,font=("calibre",13,"normal"),bg="white",justify='center')
    issuername_entry.grid(row=1,column=1)
    lb_roll=tk.Label(frame3,text="enter the roll number",width=70,font=("calibre",13,"normal"),bg="white")
    lb_roll.grid(row=3,column=0)
    roll_entry=tk.Entry(frame3,textvariable=rollnumber,width=20,font=("calibre",13,"normal"),bg="#f7f1e3")
    roll_entry.grid(row=3,column=1)
    lb_issuedate=tk.Label(frame3,text="enter the date of issue",width=70,font=("calibre",13,"normal"),bg="white")
    lb_issuedate.grid(row=5,column=0)
    issuedate_entry=tk.Entry(frame3,textvariable=issuingdate,width=20,font=('calibre',13,'bold'))
    issuedate_entry.grid(row=5,column=1)
    lb_expiry=tk.Label(frame3,text="enter the date of expiry",width=70,font=("calibre",13,"normal"),bg="white")
    lb_expiry.grid(row=9,column=0)
    expiry_entry=tk.Entry(frame3,textvariable=expiry,width=20,font=('calibre',13,'bold'))
    expiry_entry.grid(row=9,column=1)

    #button6 = Button(frame3, text="Insert Frame Stuff here",command= searchissue)







    #button6.pack(side = TOP)

def onClick():
        if len(root.winfo_children()) > 1:
             root.winfo_children()[1].destroy()

        button6 = Button(frame2, text="Insert Frame Stuff here")
    # msg = messagebox.showinfo("BUTTON CLICKED","Hello World")
button1 = Button(frame1, text="Issue Book",command = onClick1)
'''button2 = Button(frame1, text="Return Book",command = onClick)
button3 = Button(frame1, text="Add New Books",command = onClick)
button4=  Button(frame1, text="Delete Books",command = onClick)
button5 = Button(frame1, text="Replace Books",command = onClick)'''
button1.pack(fill=BOTH,side = LEFT, anchor=N)
'''button2.pack(fill=BOTH,side = LEFT, anchor=N)
button3.pack(fill=BOTH,side = LEFT, anchor=N)
button4.pack(fill=BOTH,side = LEFT, anchor=N)
button5.pack(fill=BOTH,side = LEFT, anchor=N)'''
root.mainloop()
