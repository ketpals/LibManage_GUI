import openpyxl
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import *
from tkinter import messagebox

# from finalissue import onclick1
root = Tk()
root.title("Library System")
root.geometry("1000x500")
frame1 = Frame(root, height=30 , background="bisque")
frame1.pack(fill=BOTH,side = TOP)

def onClick1():    #issue book function

    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
#the main function for isuing a book 1-it will search bookid,2-it will check availablity,3-it will issue book and open the workbook and save all the details

   # the issue function
    def availissuebook():
            bookid=bookid_entry.get()
            issuername=issuername_entry.get()
            rollnumber=roll_entry.get()
            issuingdate=issuedate_entry.get()
            expiry=expiry_entry.get()

            nospace=0
            df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
            for x in df.index:
                if df['Book ID'][x]==bookid:
                        if df['Issuer_name'][x]=="none" or df['Issuer_name'][x]=='NONE':
                            nospace=1
                            break
                        else :
                            continue

            if(nospace==1):
                    msg = messagebox.showinfo("can be issued","available")
                    df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")

                    counter = 0
                    for x in df['Book ID']:

                        counter = counter + 1

                        #if df['bookid'][x]==bookid:
                        if x == bookid:
                            #opening the excel sheet using openpyxl
                            print("counter is"+str(counter))
                            wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                            sheet = wb.active
                            i=counter +1
                            sheet.cell(row=i, column=7).value = issuername
                            print(sheet.cell(row=i, column=7).value)


                            sheet.cell(row=i,column=8).value=rollnumber
                            sheet.cell(row=i,column=9).value=issuingdate

                            sheet.cell(row=i,column=10).value=expiry

                            counter=0



                            wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                            msg=messagebox.showinfo("success","book is issued")#/*
                             #this will delete the content given on screen for refilling it

                            bookid_entry.delete(0, END)
                            issuername_entry.delete(0, END)
                            roll_entry.delete(0, END)
                            issuedate_entry.delete(0, END)
                            expiry_entry.delete(0, END)
                            messagebox.showinfo("want to issue books?","Enter the data in the blank space and press 'Submit'")

                            break
            else:
                    msg = messagebox.showinfo("cant be issued","no availability")
    #calling search function to search if book id is available,then if book found pass it to the issue book if available function
    #if book is not present pop up will come and book is unavailable to be issued or book not found

    def searching():
            bookid=bookid_entry.get()
            df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
            found=0
            for x in df.index:
                if df['Book ID'][x]==str(bookid):
                     found=1
                else:
                    continue
            if (found==1):
                msg = messagebox.showinfo("search completed","book  found")
                s2_button=tk.Button(frame3, text="issuebook if AV",command= availissuebook)
                s2_button.grid(row=0,column=13)
            else:
                msg = messagebox.showinfo("search completed","book not found")

    #frame3 =tk.Frame(root, width=400, height=500 , bg="pink")
    frame3=tk.Frame(root,width=400,height=500,background="pink")
    frame3.pack_propagate(False)
    frame3.pack(side=TOP,fill=BOTH)
    compare=tk.StringVar()
    bookid=tk.StringVar()
    issuername=tk.StringVar()
    rollnumber=tk.StringVar()
    issuingdate=tk.StringVar()
    expiry=tk.StringVar()
    lb_none=tk.Label(frame3,text="no",width=1,font=("calibre",10,"normal"),bg="white")
    lb_none.grid(row=0,column=23)

    lb_bookid=tk.Label(frame3,text="enter the book ID to be searched",width=70,font=("calibre",13,"normal"),bg="white")
    lb_bookid.grid(row=0,column=0)
    bookid_entry=tk.Entry(frame3,textvariable=bookid,width=20,font=('calibre',13,'bold'))
    bookid_entry.grid(row=0,column=1)

    sub_button=tk.Button(frame3, text="search",command= searching)
    sub_button.grid(row=0,column=14)

    lb_issuername=tk.Label(frame3,text="enter the name of issuer",width=70,font=("calibre",13,"normal"),bg="white")
    lb_issuername.grid(row=1,column=0)
    issuername_entry=tk.Entry(frame3,textvariable=issuername,width=20,font=("calibre",13,"normal"),bg="white",justify='center')
    issuername_entry.grid(row=1,column=1)
    lb_roll=tk.Label(frame3,text="enter the roll number",width=70,font=("calibre",13,"normal"),bg="white")
    lb_roll.grid(row=3,column=0)
    roll_entry=tk.Entry(frame3,textvariable=rollnumber,width=20,font=("calibre",13,"normal"),bg="#f7f1e3")
    roll_entry.grid(row=3,column=1)
    lb_issuedate=tk.Label(frame3,text="enter the date of issue",width=70,font=("calibre",13,"normal"),bg="white")
    lb_issuedate.grid(row=5,column=0)
    issuedate_entry=tk.Entry(frame3,textvariable=issuingdate,width=20,font=('calibre',13,'bold'))
    issuedate_entry.grid(row=5,column=1)
    lb_expiry=tk.Label(frame3,text="enter the date of expiry",width=70,font=("calibre",13,"normal"),bg="white")
    lb_expiry.grid(row=9,column=0)
    expiry_entry=tk.Entry(frame3,textvariable=expiry,width=20,font=('calibre',13,'bold'))
    expiry_entry.grid(row=9,column=1)

def bookRegister():       # helper function of ADD BOOK
    
    #messagebox is imported to display message boxes
    #pandas library is imported to create dataframe
    #openpyxl library is imported to write / read excel sheets

    title = Info1.get()
    author =Info2.get()
    date = Info3.get()
    bktype =Info4.get()
    bookid =Info5.get()
    
    if(title=='' or author=='' or date=='' or bktype=='' or bookid==''):
        messagebox.showwarning("Invalid","All the fields are required")
        return
        #if some or all the entries are empty, warning is shown, and it returns to the calling function
        #else the data is added to the excel sheet
    else:
        wb = openpyxl.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx") #excel sheet is loaded in the workbook
        sheet = wb.active                        
        issuer_name= "none"
        serial=newRowLoc=sheet.max_row 
        bktype=bktype.lower()
        about_book=[serial,title,author,date,bktype,bookid,issuer_name] #a list is created of all the data entries of book
        sheet.append(about_book)             # the list is added at the end of excel file retaining old data
        wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx")          #saves the excel file
    
        messagebox.showinfo('Success',"Book added successfully") #message is displayed after successful addition of data to excel sheet
        #all the previous entries are deletd after the data is saved to file
        Info5.delete(0, END) 
        Info4.delete(0, END)
        Info3.delete(0, END)
        Info2.delete(0, END)
        Info1.delete(0, END)   
def addBook():     #add book function
    
    global Info1 ,Info2,Info3, Info4,Info5,Info6
    
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
    
    frame1 = Frame(root,bg="#f7f1e3") #frame is used  for arranging the position of other widgets
    frame1.pack(expand=True,fill=BOTH) #expand is set to true so that frame1 expands to fill any space not otherwise used in frame1's parent.
    #fill=BOTH fills ( both horizontally and vertically ) any extra space allocated to it
    
    #all the variables holds a string 
    title=StringVar()
    author=StringVar()
    date=StringVar()
    bktype=StringVar()
    bookid=StringVar()
        
    #heading is given to the frame ,also headingFrame1 is used to hold the heading which is above frame1
    headingFrame1 = Frame(frame1,bg="#6bd5e8",bd=5) 
    headingFrame1.place(relx=0.25,rely=0.03,relwidth=0.5,relheight=0.12)
    headingLabel = Label(headingFrame1, text="Add Books", bg="#6bd5e8", fg='#151c52', font=('Courier',19,'bold'))
    headingLabel.place(relx=0,rely=0, relwidth=1, relheight=1)
    
    #frame to hold all the blank space and their labels 
    labelFrame = Frame(frame1,bg="#6bd5e8")
    labelFrame.place(relx=0.1,rely=0.20,relwidth=0.8,relheight=0.6)
        
    #relx and rely: Horizontal and vertical offset as a float between 0.0 and 1.0, as a fraction of the height and width of the parent widget.
    #relheight, relwidth - Height and width as a float between 0.0 and 1.0, as a fraction of the height and width of the parent widget
    #relx, rely, relheight, relwidth are used for every label and entry and also for two buttons
    # Title
    lb1 = Label(labelFrame,text="Title : ", bg="#6bd5e8", fg='black',font=("calibre",14,'normal'))
    lb1.place(relx=0.05,rely=0.10, relheight=0.08)
        
    Info1 = Entry(labelFrame,justify='center',textvariable=title,font=("calibre",14,'normal'))
    Info1.place(relx=0.4,rely=0.10, relwidth=0.56, relheight=0.09)
        
    # Book Author
    lb2 = Label(labelFrame,text="Author : ", bg="#6bd5e8", fg='black',font=("calibre",14,'normal'))
    lb2.place(relx=0.05,rely=0.25, relheight=0.08)
        
    Info2 = Entry(labelFrame,textvariable=author,font=("calibre",14,'normal'),justify='center')
    Info2.place(relx=0.4,rely=0.25, relwidth=0.56, relheight=0.09)
    
    #Addition Date
    lb3 = Label(labelFrame,text="Addition Date : ", bg="#6bd5e8", fg='black',font=("calibre",14,'normal'))
    lb3.place(relx=0.05,rely=0.40, relheight=0.08)
        
    Info3 = Entry(labelFrame,textvariable=date,font=("calibre",14,'normal'),justify='center')
    Info3.place(relx=0.4,rely=0.40, relwidth=0.56, relheight=0.09)
    
    #type of book
    lb4 = Label(labelFrame,text="Type of book : ", bg="#6bd5e8", fg='black',font=("calibre",14,'normal'))
    lb4.place(relx=0.05,rely=0.55, relheight=0.08)
        
    Info4 = Entry(labelFrame,textvariable=bktype,font=("calibre",14,'normal'),justify='center')
    Info4.place(relx=0.4,rely=0.55, relwidth=0.56, relheight=0.09)
    
    # Book ID
    lb5 = Label(labelFrame,text="Book ID : ", bg="#6bd5e8", fg='black',font=("calibre",14,'normal'))
    lb5.place(relx=0.05,rely=0.70, relheight=0.08)
        
    Info5 = Entry(labelFrame,textvariable=bookid,font=("calibre",14,'normal'),justify='center')
    Info5.place(relx=0.4,rely=0.70, relwidth=0.56, relheight=0.09)
        
    #two buttons are on frame1 
    #bg is background colour
    #fg is forground colour
    #text is displayed over the buttons
    
    #Submit Button
    SubmitBtn= Button(frame1,text="SUBMIT",bg='#d4d4d4', fg='black',command= bookRegister,font=("calibre",12,'normal'))
    SubmitBtn.place(relx=0.42,rely=0.85, relwidth=0.18,relheight=0.08)  
    #if the user clicks Submit button, bookRegister() function is called. this function adds the data to the excel sheet.  
 
def delete_bk():      #delete function 
    df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #reading from the excel file
    bookid= bookid_entry.get()     #variable for book ID
    Bookname=Bookname_entry.get()   #variable for Book Name
    if (bookid=="" or Bookname==" "):
        messagebox.showwarning("Invalid","All the fields are required")
        return
    else:
        for x in df.index:       
            if df['Book ID'][x]==bookid:    #condition checking if the book is present or not
                msg=messagebox.showinfo("BOOK IS PRESENT","CAN BE DELETED")
                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #loading the workbook
                sheet = wb.active     #making the excel file actie to work on
                a=df['serial'][x]         
                b=a+1              #logic for rows so that it gets incremented accordingly
                if(sheet.cell(row=b, column=6).value == bookid):    # checking if the bookid entered is present in bookid column
                    sheet.cell(row=b, column=1).value = "-"        #if does then replace it with "-"
                    sheet.cell(row=b, column=2).value = "-"
                    sheet.cell(row=b, column=3).value = "-"
                    sheet.cell(row=b, column=4).value = "-"
                    sheet.cell(row=b, column=5).value = "-"
                    sheet.cell(row=b, column=6).value = "-"
                    sheet.cell(row=b, column=7).value = "-"
                    sheet.cell(row=b, column=8).value = "-"
                    sheet.cell(row=b, column=9).value = "-"
                    sheet.cell(row=b, column=10).value = "-"
                    sheet.cell(row=b, column=11).value = "-"
                    sheet.cell(row=b, column=12).value = "-"  
                    
                    index_row = []
                    for i in range(1, sheet.max_row):          #for shifting the next row in place of deleted or made "-" row
                        if sheet.cell(i, 1).value == "-":
                            index_row.append(i)
                    for row_del in range(len(index_row)):
                        sheet.delete_rows(idx=index_row[row_del], amount=1)
                        index_row = list(map(lambda k: k - 1, index_row))
                    
                       #saving the excel file after execution
                    wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx") 
                break
        if df['Book ID'][x]==bookid:
                
                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  #loading the workbook
                sheet = wb.active     
                counter = 0
                for column in range(1,2):
                    column_letter = get_column_letter(column)
                for row in range(2,sheet.max_row+1):
                    counter = counter +1
                    sheet[column_letter + str(row)] = counter
                wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx") 
                msg=messagebox.showinfo("Success","Book have been deleted successfully")
                Bookname_entry.delete(0, END)
                bookid_entry.delete(0, END)
        else:
                msg=messagebox.showinfo("Book Not Present","Cannot be deleted")
                
def Delete_book():     #the main function which is called upon clicking the delete button in the menu
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
            #logic for rows so that it gets incremented accordingly
    frame1 = Frame(root,bg="#fafafa")
    frame1.pack(expand=True,fill=BOTH)
    
    headingFrame12 = Frame(frame1,bg="#f7ca86",bd=5) 
    headingFrame12.place(relx=0.22,rely=0.025,relwidth=0.55,relheight=0.10)
    headingLabel2 = Label(headingFrame12, text="Delete a Book", bg="#f7ca86", fg='#112a4f', font=('Courier',19,'bold'))
    headingLabel2.place(relx=0,rely=0, relwidth=1, relheight=1)
    bookid= StringVar()
    Bookname=StringVar()
    global bookid_entry,Bookname_entry
    
    frame2 = Frame(frame1 ,bg="#f7ca86") 
    frame2.place(relx=0.05,rely=0.20,relwidth=0.9,relheight=0.6)
    
    lb_Bookname=Label(frame2,text="Enter the book name to be deleted",font=("calibre",13,"normal"),bg="#f7ca86")
    lb_Bookname.place(relx=0.05,rely=0.30, relheight=0.08)
   
    Bookname_entry=Entry(frame2,textvariable=Bookname,font=('calibre',13,'bold'))
    Bookname_entry.place(relx=0.42,rely=0.30, relwidth=0.50, relheight=0.09)
        
    lb_bookid=Label(frame2,text="Enter the Book ID",font=("calibre",13,"normal"),bg="#f7ca86")
    lb_bookid.place(relx=0.05,rely=0.60, relheight=0.08)
    
    bookid_entry=Entry(frame2,textvariable=bookid,font=('calibre',13,'bold'))
    bookid_entry.place(relx=0.42,rely=0.60, relwidth=0.50, relheight=0.09)
    
    sub_button=Button(frame1,text="Delete",command= delete_bk,bg='#d4d4d4', fg='black',font=("calibre",12,'normal'))
    sub_button.place(relx=0.40,rely=0.85, relwidth=0.18,relheight=0.08)


def rturn():      #Function to return a book after being issued
    file = pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
    serial_no=str(sne.get())
    if(len(file.index[file['Book ID'] == serial_no].values)==0):
        messagebox.showinfo("Error","Serial Number Not Exists")
        return
    file.loc[file['Book ID']==serial_no,'Issuer_name'] = "none"
    file.loc[file['Book ID']==serial_no,'Issue Date'] = "none"
    file.loc[file['Book ID']==serial_no,'Roll no'] = "none"
    file.loc[file['Book ID']==serial_no,'Return date'] = "none"
    file.loc[file['Book ID']==serial_no,'Expiry'] = "none"
    file.to_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx", index=False)
    messagebox.showinfo("Done","Done Returning")

def onClick():
    global sne,rne,ne,bne
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
    frame = Frame(root)
    frame.pack()
    snl = Label(frame, text = "Book ID")
    sne = Entry(frame)
    snl.grid(row = 0, column = 0, sticky = W, pady = 2)
    sne.grid(row = 0, column = 1, sticky = W, pady = 2)

    ok = Button(frame, text="OK",command = rturn)
    ok.grid(row = 1, column = 0, columnspan = 4 ,sticky = W, pady = 2)

def onClick3():
#the main function for isuing a book 1-it will search bookid,2-it will check availablity,3-it will issue book and open the workbook and save all the details 
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
   
    frame4 = Frame(root , width=400, height=500 , bg="pink")
    frame4.pack_propagate(False)
    frame4.pack(side=TOP,fill=BOTH)
       
    bookid=tk.StringVar()
    authorname=tk.StringVar()
    types=tk.StringVar()
    bookname=tk.StringVar()
    additiondates=tk.StringVar() 
     
    def look():   
        bookid=bookid_entry.get()
        bookname=bookname_entry.get()
        authorname=authorname_entry.get()
        additiondates=additiondates_entry.get()
        df=pd.read_excel(r"C:\Users\Admin\Downloads\Library Books.xlsx")
        found=0 
        for x in df.index:
            if df['Book ID'][x]==bookid:
                found=1
                if df['Issuer_name'][x]=="none" or df['Issuer_name'][x]=="None":
                        c=0
                        for y in df['Book ID']:
                            c = c + 1
                            if y ==bookid:
                                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  
                                sheet = wb.active 
                                i=c +1
                                sheet.cell(row=i, column=2).value =bookname
                                sheet.cell(row=i,column=3).value=authorname          
                                sheet.cell(row=i,column=4).value=additiondates 
                                sheet.cell(row=i,column=13).value="1"
                                wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                                msg=messagebox.showinfo("success","book is replaced")
                                break
                else:
                    msg=messagebox.showinfo("failed","book is issued plz wait sometime")
        else:
            pass
        if (found==0):   
                    msg = messagebox.showinfo("cant replace","book is not in library ")   
    lb_id=tk.Label(frame4,text="enter the book ID",width=70,font=("calibre",13,"normal"),bg="white")
    lb_id.grid(row=0,column=0)
    bookid_entry=tk.Entry(frame4,textvariable=types,width=20,font=('calibre',13,'bold'))
    bookid_entry.grid(row=0,column=1)
    lb_bookname=tk.Label(frame4,text="change the book name",width=70,font=("calibre",13,"normal"),bg="white")
    lb_bookname.grid(row=1,column=0)
    bookname_entry=tk.Entry(frame4,textvariable=bookname,width=20,font=("calibre",13,"normal"),bg="white",justify='center')
    bookname_entry.grid(row=1,column=1)
    lb_authorname=tk.Label(frame4,text="enter the authorname",width=70,font=("calibre",13,"normal"),bg="white")
    lb_authorname.grid(row=3,column=0)
    authorname_entry=tk.Entry(frame4,textvariable=authorname,width=20,font=("calibre",13,"normal"),bg="#f7f1e3")
    authorname_entry.grid(row=3,column=1)
    lb_additiondates=tk.Label(frame4,text="enter the date of addition of book",width=70,font=("calibre",13,"normal"),bg="white")
    lb_additiondates.grid(row=5,column=0)
    additiondates_entry=tk.Entry(frame4,textvariable=additiondates,width=20,font=('calibre',13,'bold'))
    additiondates_entry.grid(row=5,column=1)
        
    sub_button=tk.Button(frame4, text="ok",command= look)
    sub_button.grid(row=0,column=14)

button1 = Button(frame1, text="Issue Book",command = onClick1)
button2 = Button(frame1, text="Return Book",command = onClick)
button3 = Button(frame1, text="Add New Books",command = addBook)
button4 = Button(frame1, text="Delete Books",command = Delete_book)
button5 = Button(frame1, text="Replace Books",command = onClick3)

button1.pack(fill=BOTH,side = LEFT, anchor=N)
button2.pack(fill=BOTH,side = LEFT, anchor=N)
button3.pack(fill=BOTH,side = LEFT, anchor=N)
button4.pack(fill=BOTH,side = LEFT, anchor=N)
button5.pack(fill=BOTH,side = LEFT , anchor=N)


root.mainloop()