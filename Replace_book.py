import pandas as pd
import openpyxl as op
import tkinter as tk
from tkinter import *
from tkinter import messagebox
root = Tk()
root.geometry("400x400")
frame1 = Frame(root, height=30 , background="bisque")
frame1.pack_propagate(False)
frame1.pack(fill=BOTH)
def onClick3():
#the main function for isuing a book 1-it will search bookid,2-it will check availablity,3-it will issue book and open the workbook and save all the details 
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()
   
    frame3 = Frame(root , width=400, height=500 , background="pink")
    frame3.pack_propagate(False)
    frame3.pack(side=TOP,fill=BOTH)
       
    bookid=tk.StringVar()
    authorname=tk.StringVar()
    types=tk.StringVar()
    bookname=tk.StringVar()
    additiondates=tk.StringVar() 
     
    def look():   
        bookid=bookid_entry.get()
        bookname=bookname_entry.get()
        authorname=authorname_entry.get()
        additiondates=additiondates_entry.get()
        df=pd.read_excel(r"Library Books.xlsx")
        found=0 
        for x in df.index:
            if df['Book ID'][x]==bookid:
                
                found=1
                if df['Issuer_name'][x]=="none" or df['Issuer_name'][x]=="None" :
                        c=0
                        for y in df['Book ID']:
                            c = c + 1
                            if y ==bookid:
                                wb = op.load_workbook(r"C:\Users\Admin\Downloads\Library Books.xlsx")  
                                sheet = wb.active 
                                i=c +1
                                sheet.cell(row=i, column=2).value =bookname
                                sheet.cell(row=i,column=3).value=authorname          
                                sheet.cell(row=i,column=4).value=additiondates 
                                sheet.cell(row=i,column=13).value="1"
                                wb.save(r"C:\Users\Admin\Downloads\Library Books.xlsx")
                                msg=messagebox.showinfo("success","book is replaced")
                                break
                else:
                    msg=messagebox.showinfo("failed","book is issued plz wait sometime")
        else:
            pass
        if (found==0):   
                    msg = messagebox.showinfo("cant replace","book is not in library ")   
    lb_id=tk.Label(frame3,text="enter the book ID",width=70,font=("calibre",13,"normal"),bg="white")
    lb_id.grid(row=0,column=0)
    bookid_entry=tk.Entry(frame3,textvariable=types,width=20,font=('calibre',13,'bold'))
    bookid_entry.grid(row=0,column=1)
    lb_bookname=tk.Label(frame3,text="change the book name",width=70,font=("calibre",13,"normal"),bg="white")
    lb_bookname.grid(row=1,column=0)
    bookname_entry=tk.Entry(frame3,textvariable=bookname,width=20,font=("calibre",13,"normal"),bg="white",justify='center')
    bookname_entry.grid(row=1,column=1)
    lb_authorname=tk.Label(frame3,text="enter the authorname",width=70,font=("calibre",13,"normal"),bg="white")
    lb_authorname.grid(row=3,column=0)
    authorname_entry=tk.Entry(frame3,textvariable=authorname,width=20,font=("calibre",13,"normal"),bg="#f7f1e3")
    authorname_entry.grid(row=3,column=1)
    lb_additiondates=tk.Label(frame3,text="enter the date of addition of book",width=70,font=("calibre",13,"normal"),bg="white")
    lb_additiondates.grid(row=5,column=0)
    additiondates_entry=tk.Entry(frame3,textvariable=additiondates,width=20,font=('calibre',13,'bold'))
    additiondates_entry.grid(row=5,column=1)
        
    sub_button=tk.Button(frame3, text="ok",command= look)
    sub_button.grid(row=0,column=14)
def onClick():   
    if len(root.winfo_children()) > 1:
        root.winfo_children()[1].destroy()    
    # for widget in root.winfo_children():
        # print(widget)
            # widget.destroy()
'''button1 = Button(frame1, text="Issue Book",command = onClick)
button2 = Button(frame1, text="Return Book",command = onClick)
button3 = Button(frame1, text="Add New Books",command = onClick)'''
button4 = Button(frame1, text="Replace Books",command = onClick3)
'''button1.pack(fill=BOTH,side = LEFT, anchor=N)
button2.pack(fill=BOTH,side = LEFT, anchor=N)
button3.pack(fill=BOTH,side = LEFT, anchor=N)'''
button4.pack(fill=BOTH,side = LEFT, anchor=N)
root.mainloop()